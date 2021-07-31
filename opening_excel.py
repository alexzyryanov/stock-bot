import openpyxl
import sqlite3


def create(user_id):
    file = openpyxl.load_workbook(f"{user_id}.xlsx")
    sheet = file['Сделки']

    line = sheet.max_row
    while line != 1:
        ticker = sheet.cell(row=line, column=5)
        operation = sheet.cell(row=line, column=8)
        count = sheet.cell(row=line, column=9)
        prise = sheet.cell(row=line, column=17)

        add_in_db(user_id, ticker.value, operation.value, count.value, float('{:.2f}'.format(prise.value)))
        line -= 1


def add_in_db(user_id, ticker, operation, count, prise):
    connect_db = sqlite3.connect("bot.db")
    cursor = connect_db.cursor()
    cursor.execute(f""" CREATE TABLE IF NOT EXISTS users (user INT, ticker TEXT, count INT, cost REAL) """)
    connect_db.commit()

    cursor.execute(f""" SELECT ticker FROM users WHERE ticker = "{ticker}" AND user = "{user_id}" """)
    check = cursor.fetchone()
    if check is None:
        cursor.execute(f""" INSERT INTO users VALUES (?, ?, ?, ?) """, (user_id, ticker, count, prise))
        connect_db.commit()
    else:
        if operation == "Покупка":
            cursor.execute(f""" SELECT * FROM users WHERE ticker = "{ticker}" AND user = "{user_id}" """)
            old = cursor.fetchone()
            cursor.execute(f""" UPDATE users 
                                SET count = "{old[2] + count}", cost = "{old[3] + prise}" 
                                WHERE ticker = "{ticker}" AND user = "{user_id}" """)
            connect_db.commit()
        elif operation == "Продажа":
            cursor.execute(f""" SELECT * FROM users WHERE ticker = "{ticker}" AND user = "{user_id}" """)
            old = cursor.fetchone()
            if old[2] - count == 0:
                cursor.execute(f""" DELETE FROM users WHERE ticker = "{ticker}" AND user = "{user_id}" """)
                connect_db.commit()
            else:
                cursor.execute(f""" UPDATE users 
                                    SET count = "{old[2] - count}", cost = "{old[3] - prise}" 
                                    WHERE ticker = "{ticker}" AND user = "{user_id}" """)
                connect_db.commit()

    cursor.close()
    connect_db.close()


def check(user_id):
    connect_db = sqlite3.connect("bot.db")
    cursor = connect_db.cursor()
    user_prise = 0
    now_prise = 0
    cursor.execute(f""" SELECT ticker, count, cost FROM users WHERE user = "{user_id}" """)
    data = cursor.fetchall()
    for i in data:
        user_prise += i[2]
        cursor.execute(f""" SELECT cost FROM ru WHERE ticker = "{i[0]}" """)
        t = cursor.fetchone()
        now_prise += float(t[0]) * int(i[1])
    cursor.close()
    connect_db.close()
    return float('{:.2f}'.format(now_prise - user_prise))


def clean(user_id):
    connect_db = sqlite3.connect("bot.db")
    cursor = connect_db.cursor()
    cursor.execute(f""" DELETE FROM users WHERE user = "{user_id}" """)
    connect_db.commit()
    cursor.close()
    connect_db.close()
